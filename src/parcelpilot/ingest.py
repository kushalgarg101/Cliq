from __future__ import annotations

import hashlib
import io
import logging
import os
import zipfile
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from pypdf import PdfReader

from .db import connect

DOCUMENTS = {
    "01_Support_Policy_v3_CURRENT.pdf": ("support_policy", "current", 70, None),
    "02_Support_Policy_v2_DEPRECATED.pdf": ("support_policy", "deprecated", 0, None),
    "03_Cancellation_and_Service_Credit_SOP_v4.pdf": ("cancellation_sop", "current", 80, None),
    "04_Product_Operations_Guide_and_Known_Issues.pdf": ("product_guide", "current", 60, None),
    "05_Northstar_Logistics_Enterprise_Agreement.pdf": ("agreement", "current", 100, "ACCT-001"),
    "06_LumenWorks_Service_Agreement.pdf": ("agreement", "current", 100, "ACCT-002"),
}
logger = logging.getLogger(__name__)

def ingest(zip_path: Path, destination: Path) -> None:
    if not zip_path.exists():
        raise FileNotFoundError(f"Candidate data pack not found: {zip_path}")
    if destination.exists():
        raise FileExistsError(f"Refusing to overwrite an active generated database: {destination}")
    temporary_destination = destination.with_name(f".{destination.name}.{uuid4().hex}.tmp")
    try:
        with zipfile.ZipFile(zip_path) as archive, connect(temporary_destination) as db:
            names = archive.namelist()
            workbook_matches = [name for name in names if Path(name).suffix.lower() == ".xlsx"]
            if len(workbook_matches) != 1:
                raise ValueError("Candidate pack must contain exactly one workbook.")
            document_entries = {}
            for filename in DOCUMENTS:
                matches = [name for name in names if Path(name).name == filename]
                if len(matches) != 1:
                    raise ValueError(f"Candidate pack is missing or duplicates {filename}.")
                document_entries[filename] = matches[0]
            db.executescript(
            """
            CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT NOT NULL);
            CREATE TABLE documents (id INTEGER PRIMARY KEY, filename TEXT UNIQUE, category TEXT, status TEXT,
              authority INTEGER, account_id TEXT, text TEXT);
            CREATE TABLE chunks (id INTEGER PRIMARY KEY, document_id INTEGER, page INTEGER, text TEXT,
              FOREIGN KEY(document_id) REFERENCES documents(id));
            CREATE TABLE accounts (account_id TEXT PRIMARY KEY, account_name TEXT, plan TEXT, status TEXT,
              csm TEXT, contract_file TEXT, premium_support TEXT, notes TEXT);
            CREATE TABLE orders (order_id TEXT PRIMARY KEY, account_id TEXT, carrier TEXT, status TEXT,
              booked_at TEXT, pickup_window_start TEXT, pickup_window_end TEXT, pickup_actual_at TEXT,
              shipment_fee_inr TEXT, carrier_fault TEXT, customer_fault TEXT, cancellation_requested_at TEXT, notes TEXT);
            CREATE TABLE tickets (ticket_id TEXT PRIMARY KEY, account_id TEXT, created_at TEXT, status TEXT,
              subject TEXT, description TEXT, channel TEXT, assigned_to TEXT, last_customer_message_at TEXT,
              historical_resolution TEXT);
            """
            )
            workbook = load_workbook(io.BytesIO(archive.read(workbook_matches[0])), data_only=True)
            for sheet_name in ("README", "accounts", "orders", "tickets"):
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Candidate workbook is missing the {sheet_name} sheet.")
            snapshot = workbook["README"]["B2"].value
            if not snapshot:
                raise ValueError("Candidate workbook does not define a dataset snapshot time.")
            db.execute("INSERT INTO meta VALUES (?, ?)", ("snapshot_at", str(snapshot)))
            db.execute("INSERT INTO meta VALUES (?, ?)", ("source_fingerprint", _fingerprint(zip_path)))
            for sheet_name in ("accounts", "orders", "tickets"):
                worksheet = workbook[sheet_name]
                headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
                if not all(headers):
                    raise ValueError(f"Candidate workbook has an invalid header row in {sheet_name}.")
                placeholders = ",".join("?" for _ in headers)
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    db.execute(f"INSERT INTO {sheet_name} VALUES ({placeholders})", tuple("" if value is None else str(value) for value in row))
            for filename, metadata in DOCUMENTS.items():
                category, status, authority, account_id = metadata
                reader = PdfReader(io.BytesIO(archive.read(document_entries[filename])))
                pages = [page.extract_text() or "" for page in reader.pages]
                cursor = db.execute(
                    "INSERT INTO documents(filename, category, status, authority, account_id, text) VALUES (?, ?, ?, ?, ?, ?)",
                    (filename, category, status, authority, account_id, "\n".join(pages)),
                )
                for page_number, page_text in enumerate(pages, start=1):
                    for chunk in [page_text[index:index + 1100] for index in range(0, len(page_text), 900)]:
                        db.execute("INSERT INTO chunks(document_id, page, text) VALUES (?, ?, ?)", (cursor.lastrowid, page_number, chunk))
        os.replace(temporary_destination, destination)
    finally:
        try:
            temporary_destination.unlink(missing_ok=True)
        except OSError:
            logger.warning("Could not remove incomplete generated source database %s.", temporary_destination)


def ensure_ingested(zip_path: Path | None, destination: Path) -> bool:
    if not zip_path or not zip_path.is_file():
        logger.error("Candidate data pack is not configured or cannot be read.")
        return False
    if destination.exists():
        try:
            with connect(destination) as db:
                existing = db.execute("SELECT value FROM meta WHERE key='source_fingerprint'").fetchone()
            if not existing:
                logger.warning("Generated source database has no fingerprint; rebuild it from the candidate pack before serving data.")
                return False
            if existing[0] == _fingerprint(zip_path):
                return True
        except Exception:
            logger.exception("Could not inspect the generated source database.")
            return False
        logger.warning("Source data changed; rebuild var/source.db before serving data.")
        return False
    try:
        ingest(zip_path, destination)
        return True
    except Exception:
        logger.exception("Could not ingest the candidate data pack.")
        return False


def _fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()
